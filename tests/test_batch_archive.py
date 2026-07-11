"""Tests for `app.core.batch_archive` — P18-1 file-backed archive storage.

Covers: batch/job/verdict CRUD round-trips; crash recovery (a leftover
partial-write temp file must never surface as a real record); the [EXPERT]
verdict red line (blank verdict_text/reviewer never constructs, "not yet
recorded" is representable only by absence); xlsx export column separation
(machine columns vs [EXPERT] columns on physically different sheets, absent
verdicts leave no row rather than a placeholder).
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from pydantic import ValidationError

from app.core.batch_archive import (
    BatchArchive,
    BatchArchiveError,
    BatchJobFailure,
    BatchJobRecord,
    ExpertVerdict,
    build_batch_workbook,
)

_TARGETS = [
    {"scenario": "smartphone-wide", "efl_mm": 3.8, "fnum": 2.0, "fov_deg": 80.0, "image_height_mm": 3.6},
    {"scenario": "smartphone-wide", "efl_mm": 4.8, "fnum": 1.8, "fov_deg": 79.0, "image_height_mm": 4.4},
]


def _make_job(
    *,
    batch_id: str,
    job_id: str = "job-0000",
    target_index: int = 0,
    status: str = "succeeded",
    failure: BatchJobFailure | None = None,
) -> BatchJobRecord:
    return BatchJobRecord(
        job_id=job_id,
        batch_id=batch_id,
        target_index=target_index,
        target_label="target-0000",
        target_spec=_TARGETS[0],
        status=status,
        created_at="2026-07-11T00:00:00+00:00",
        updated_at="2026-07-11T00:00:01+00:00",
        result_summary={"candidate_count": 4, "ranked_count": 3, "withheld_count": 1},
        candidate_set_pointer=None,
        artifact_dir=None,
        failure=failure,
    )


# ---------------------------------------------------------------------------
# Batch / targets CRUD
# ---------------------------------------------------------------------------


def test_create_and_get_batch_round_trips(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    created = archive.create_batch(target_source="unit-test", targets=_TARGETS, engine="fake")

    fetched = archive.get_batch(created.batch_id)
    assert fetched == created
    assert fetched.target_count == len(_TARGETS)
    assert fetched.status == "running"


def test_create_batch_freezes_targets_to_disk(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    created = archive.create_batch(target_source="unit-test", targets=_TARGETS, engine="fake")

    assert archive.get_targets(created.batch_id) == _TARGETS


def test_create_batch_duplicate_id_raises(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    archive.create_batch(target_source="x", targets=_TARGETS, engine="fake", batch_id="dup")
    with pytest.raises(BatchArchiveError):
        archive.create_batch(target_source="x", targets=_TARGETS, engine="fake", batch_id="dup")


def test_get_unknown_batch_raises(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    with pytest.raises(BatchArchiveError):
        archive.get_batch("nope")


def test_list_batches_sorted_newest_first(tmp_path: Path, monkeypatch):
    import app.core.batch_archive as batch_archive_module

    archive = BatchArchive(root=tmp_path)
    monkeypatch.setattr(batch_archive_module, "_utc_now_iso", lambda: "2026-01-01T00:00:00+00:00")
    b1 = archive.create_batch(target_source="a", targets=_TARGETS, engine="fake", batch_id="aaa")
    monkeypatch.setattr(batch_archive_module, "_utc_now_iso", lambda: "2026-02-01T00:00:00+00:00")
    b2 = archive.create_batch(target_source="b", targets=_TARGETS, engine="fake", batch_id="bbb")

    listed = archive.list_batches()
    assert [b.batch_id for b in listed] == [b2.batch_id, b1.batch_id]


def test_update_batch_changes_status_and_bumps_updated_at(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    created = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")

    updated = archive.update_batch(created.batch_id, status="completed")
    assert updated.status == "completed"
    assert updated.updated_at >= created.updated_at
    assert archive.get_batch(created.batch_id).status == "completed"


# ---------------------------------------------------------------------------
# Job CRUD
# ---------------------------------------------------------------------------


def test_put_and_get_job_round_trips(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    job = _make_job(batch_id=batch.batch_id)

    archive.put_job(job)
    fetched = archive.get_job(batch.batch_id, job.job_id)
    assert fetched == job


def test_put_job_is_idempotent_overwrite(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id, status="queued"))
    archive.put_job(_make_job(batch_id=batch.batch_id, status="succeeded"))

    assert archive.get_job(batch.batch_id, "job-0000").status == "succeeded"


def test_list_jobs_sorted_by_target_index(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id, job_id="job-0001", target_index=1))
    archive.put_job(_make_job(batch_id=batch.batch_id, job_id="job-0000", target_index=0))

    jobs = archive.list_jobs(batch.batch_id)
    assert [j.job_id for j in jobs] == ["job-0000", "job-0001"]


def test_list_jobs_on_batch_with_no_jobs_dir_returns_empty(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    assert archive.list_jobs(batch.batch_id) == []


def test_job_record_status_failed_requires_failure_category():
    with pytest.raises(ValidationError):
        BatchJobRecord(
            job_id="job-0000",
            batch_id="b",
            target_index=0,
            target_label="t",
            target_spec={},
            status="failed",
            created_at="x",
            updated_at="x",
            failure=None,
        )


def test_job_record_non_failed_status_rejects_failure_field():
    with pytest.raises(ValidationError):
        BatchJobRecord(
            job_id="job-0000",
            batch_id="b",
            target_index=0,
            target_label="t",
            target_spec={},
            status="succeeded",
            created_at="x",
            updated_at="x",
            failure=BatchJobFailure(category="engine", message="boom"),
        )


# ---------------------------------------------------------------------------
# Crash recovery: a leftover partial-write temp file must never surface.
# ---------------------------------------------------------------------------


def test_crash_mid_write_leftover_tmp_file_ignored_on_reload(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id, job_id="job-0000"))

    # Simulate a process killed mid atomic-write: a `.tmp-<hex>` sibling with
    # truncated/garbage JSON, matching what `_atomic_write_json` would leave
    # behind if the process died between `write_text` and `os.replace`.
    jobs_dir = tmp_path / batch.batch_id / "jobs"
    leftover = jobs_dir / "job-0001.json.tmp-deadbeef"
    leftover.write_text('{"job_id": "job-0001", "batch_i', encoding="utf-8")

    # A brand-new BatchArchive instance (simulating process restart) must
    # reconstruct exactly the one complete job, ignoring the leftover.
    reloaded = BatchArchive(root=tmp_path)
    jobs = reloaded.list_jobs(batch.batch_id)
    assert [j.job_id for j in jobs] == ["job-0000"]
    assert reloaded.get_batch(batch.batch_id) == batch


def test_crash_mid_write_corrupt_batch_json_skipped_in_list(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    good = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake", batch_id="good-batch")

    corrupt_dir = tmp_path / "corrupt-batch"
    corrupt_dir.mkdir()
    (corrupt_dir / "batch.json").write_text('{"batch_id": "corrupt-batch", "created_a', encoding="utf-8")

    reloaded = BatchArchive(root=tmp_path)
    listed = reloaded.list_batches()
    assert [b.batch_id for b in listed] == [good.batch_id]


# ---------------------------------------------------------------------------
# [EXPERT] verdicts — the red line
# ---------------------------------------------------------------------------


def test_verdict_blank_text_never_constructs():
    with pytest.raises(ValidationError):
        ExpertVerdict(
            job_id="job-0000",
            candidate_key="cand-a",
            verdict_text="   ",
            reviewer="张三",
            recorded_at="2026-07-11T00:00:00+00:00",
        )


def test_verdict_blank_reviewer_never_constructs():
    with pytest.raises(ValidationError):
        ExpertVerdict(
            job_id="job-0000",
            candidate_key="cand-a",
            verdict_text="值得细看",
            reviewer="",
            recorded_at="2026-07-11T00:00:00+00:00",
        )


def test_get_verdict_absent_returns_none_not_placeholder(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id))

    assert archive.get_verdict(batch.batch_id, "job-0000", "cand-a") is None


def test_put_and_get_verdict_round_trips(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id))
    verdict = ExpertVerdict(
        job_id="job-0000",
        candidate_key="cand-a",
        verdict_text="值得细看，几何合理",
        reviewer="张三",
        recorded_at="2026-07-11T00:00:00+00:00",
        note="备注",
    )

    archive.put_verdict(verdict, batch_id=batch.batch_id)
    fetched = archive.get_verdict(batch.batch_id, "job-0000", "cand-a")
    assert fetched == verdict


def test_put_verdict_unknown_job_raises(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    verdict = ExpertVerdict(
        job_id="job-nope",
        candidate_key="cand-a",
        verdict_text="值得细看",
        reviewer="张三",
        recorded_at="2026-07-11T00:00:00+00:00",
    )
    with pytest.raises(BatchArchiveError):
        archive.put_verdict(verdict, batch_id=batch.batch_id)


def test_list_verdicts_filters_by_job_id(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id, job_id="job-0000"))
    archive.put_job(_make_job(batch_id=batch.batch_id, job_id="job-0001", target_index=1))
    archive.put_verdict(
        ExpertVerdict(
            job_id="job-0000",
            candidate_key="cand-a",
            verdict_text="v1",
            reviewer="r1",
            recorded_at="t1",
        ),
        batch_id=batch.batch_id,
    )
    archive.put_verdict(
        ExpertVerdict(
            job_id="job-0001",
            candidate_key="cand-b",
            verdict_text="v2",
            reviewer="r2",
            recorded_at="t2",
        ),
        batch_id=batch.batch_id,
    )

    all_verdicts = archive.list_verdicts(batch.batch_id)
    assert len(all_verdicts) == 2
    only_job0 = archive.list_verdicts(batch.batch_id, job_id="job-0000")
    assert [v.candidate_key for v in only_job0] == ["cand-a"]


def test_verdict_resubmission_overwrites(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id))
    archive.put_verdict(
        ExpertVerdict(
            job_id="job-0000", candidate_key="cand-a", verdict_text="first",
            reviewer="r1", recorded_at="t1",
        ),
        batch_id=batch.batch_id,
    )
    archive.put_verdict(
        ExpertVerdict(
            job_id="job-0000", candidate_key="cand-a", verdict_text="revised",
            reviewer="r2", recorded_at="t2",
        ),
        batch_id=batch.batch_id,
    )

    fetched = archive.get_verdict(batch.batch_id, "job-0000", "cand-a")
    assert fetched is not None
    assert fetched.verdict_text == "revised"
    assert len(archive.list_verdicts(batch.batch_id)) == 1


# ---------------------------------------------------------------------------
# xlsx export — column separation
# ---------------------------------------------------------------------------


def test_workbook_has_three_sheets_summary_jobs_verdicts(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    job_ok = _make_job(batch_id=batch.batch_id, job_id="job-0000", status="succeeded")
    job_failed = _make_job(
        batch_id=batch.batch_id,
        job_id="job-0001",
        target_index=1,
        status="failed",
        failure=BatchJobFailure(category="engine", message="0 candidates produced"),
    )
    archive.put_job(job_ok)
    archive.put_job(job_failed)
    verdict = ExpertVerdict(
        job_id="job-0000", candidate_key="cand-a", verdict_text="值得细看",
        reviewer="张三", recorded_at="2026-07-11T00:00:00+00:00",
    )
    archive.put_verdict(verdict, batch_id=batch.batch_id)

    workbook_bytes = build_batch_workbook(
        batch, archive.list_jobs(batch.batch_id), archive.list_verdicts(batch.batch_id)
    )
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    assert wb.sheetnames == ["Summary", "Jobs", "Expert verdicts"]


def test_workbook_jobs_sheet_has_no_expert_columns(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id))

    workbook_bytes = build_batch_workbook(batch, archive.list_jobs(batch.batch_id), [])
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    jobs_ws = wb["Jobs"]
    header = [c.value for c in next(jobs_ws.iter_rows(min_row=1, max_row=1))]
    assert not any("verdict" in str(h).lower() or "expert" in str(h).lower() for h in header)


def test_workbook_verdicts_sheet_omits_row_for_unreviewed_candidate(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")
    archive.put_job(_make_job(batch_id=batch.batch_id))
    # Zero verdicts recorded for this job's candidates.

    workbook_bytes = build_batch_workbook(batch, archive.list_jobs(batch.batch_id), [])
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    verdicts_ws = wb["Expert verdicts"]
    rows = list(verdicts_ws.iter_rows(min_row=2, values_only=True))
    assert rows == []  # header only, no placeholder row


def test_workbook_summary_carries_non_pass_fail_banner(tmp_path: Path):
    archive = BatchArchive(root=tmp_path)
    batch = archive.create_batch(target_source="x", targets=_TARGETS, engine="fake")

    workbook_bytes = build_batch_workbook(batch, [], [])
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    summary_ws = wb["Summary"]
    all_text = " ".join(
        str(c.value) for row in summary_ws.iter_rows() for c in row if c.value is not None
    )
    assert "不代为判定" in all_text
