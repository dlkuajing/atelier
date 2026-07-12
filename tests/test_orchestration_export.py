"""Tests for `app.core.orchestration.export` — P17 sub-item 2 spec-sheet
export (xlsx workbook + per-candidate ZMX/.seq bundle zip).

Covers: workbook has the two expected sheets with values sourced from the
same `CandidateSet` the page renders (no verdict wording anywhere); the zip
bundle resolves a real ZMX for a retrieved (Mode1) candidate and fails
closed with an honest README when the payload's ZMX isn't on disk (the
known Mode3 temp-directory gap); the Mode3 reproduction `.seq` reconstructs
when provenance is complete and fails closed otherwise.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import os
import zipfile
from pathlib import Path

import openpyxl
import pytest

from app.core.case_library import load_case_library
from app.core.engines.stagec_field import (
    StageCFieldEvidence,
    build_stagec_machine_evidence,
    build_stagec_machine_readback,
    reconstruct_image_fields,
    resolve_field_target,
)
from app.core.optical_sample import OpticalSampleData
from app.core.orchestration.candidate import (
    CandidateSet,
    CandidateSetSummary,
    FnumAcceptedFinalEvidence,
    FnumLadderEvidence,
    GeneratedCandidate,
    GenerationMode,
    ImageQualityMetrics,
    ManufacturabilityProxy,
    MetricValue,
    OpticalExtras,
    RankResult,
    ScorecardRow,
    ScoredCandidate,
    TargetDeviation,
    TargetSpec,
)
from app.core.orchestration.export import (
    _resolve_seed_zmx_path,
    build_candidate_bundle_zip,
    build_candidate_set_workbook,
)
from app.core.zmx_ingest import ZMX_AMMO_DIR

# ---------------------------------------------------------------------------
# Fixtures / helpers (mirrors tests/test_orchestration_candidate.py)
# ---------------------------------------------------------------------------


def _first_case_with_metadata() -> OpticalSampleData:
    for case in load_case_library():
        if case.metadata is not None:
            return case
    raise AssertionError("case library has no case with metadata")


_CASE = _first_case_with_metadata()


def _metric(value: float | None = None) -> MetricValue:
    if value is None:
        return MetricValue(value=None, status="unavailable")
    return MetricValue(value=value, status="available")


def _image_quality() -> ImageQualityMetrics:
    return ImageQualityMetrics(
        mtf_sag=_metric(0.62),
        mtf_tan=_metric(0.58),
        diffraction_cutoff_lp_per_mm=_metric(810.0),
        rms_spot_radius_max_um=_metric(3.1),
        rms_spot_radius_mean_um=_metric(1.9),
        min_strehl_ratio=_metric(0.71),
        rms_wavefront_error_waves=_metric(0.08),
        field_curvature_tangential_delta_mm=_metric(0.03),
        field_curvature_sagittal_delta_mm=_metric(-0.02),
        max_distortion_pct=_metric(1.4),
        relative_illumination=_metric(),
    )


def _manufacturability() -> ManufacturabilityProxy:
    return ManufacturabilityProxy(
        total_track_mm=_CASE.paraxial.total_track_mm,
        n_pieces=5,
        has_special_glass=False,
        aspheric_term_count=_metric(),
        aspheric_surface_count=_metric(),
        chief_ray_angle_deg=_metric(28.4),
    )


def _deviations(mode: GenerationMode) -> list[TargetDeviation]:
    conv = {"efl", "fnum"} if mode is GenerationMode.TARGET_CONVERGED else set()
    rows = [
        ("efl", "exact", 3.8, 3.79, 0.01, 0.01 / 3.8),
        ("fov", "exact", 78.0, 76.5, 1.5, 1.5 / 78.0),
        ("fnum", "exact", 1.9, 1.95, 0.05, 0.05 / 1.9),
        ("imh", "exact", 3.2, 3.18, 0.02, 0.02 / 3.2),
    ]
    deviations = [
        TargetDeviation(
            field=field,
            constraint_kind=kind,
            target=target,
            achieved=achieved,
            violation=violation,
            rel_violation=rel,
            converged_toward_target=field in conv,
        )
        for field, kind, target, achieved, violation, rel in rows
    ]
    deviations.append(
        TargetDeviation(
            field="ttl",
            constraint_kind="unconstrained",
            target=None,
            achieved=_CASE.paraxial.total_track_mm,
            violation=0.0,
            rel_violation=None,
            converged_toward_target=False,
        )
    )
    return deviations


def _ranked_result() -> RankResult:
    return RankResult(score=0.812, status="ranked", coverage_pct=1.0, missing_metrics=[])


def _target_spec() -> TargetSpec:
    assert _CASE.metadata is not None
    return TargetSpec(
        scenario=_CASE.metadata.scenario,
        efl_mm=3.8,
        fov_deg=78.0,
        fnum=1.9,
        image_height_mm=3.2,
        n_elements=6,
    )


def _retrieved_candidate() -> ScoredCandidate:
    """A real Mode1 candidate whose payload IS a real case — its
    `metadata.source_zmx` genuinely exists under `ZMX_AMMO_DIR` (Mode1 never
    touches a temp directory), so this is representative of production, not
    a shortcut."""
    assert _CASE.metadata is not None
    candidate_id = f"{_CASE.metadata.case_id}::best_match"
    generated = GeneratedCandidate(
        candidate_id=candidate_id,
        mode=GenerationMode.RETRIEVED,
        source_case_id=_CASE.metadata.case_id,
        payload=_CASE,
        optical_extras=OpticalExtras(ri_by_field={"0.0": _metric(0.9)}),
        generation_notes=["检索最近邻 seed，未朝 target 优化", "role=best_match"],
    )
    scorecard = ScorecardRow(
        candidate_id=candidate_id,
        mode=GenerationMode.RETRIEVED,
        target_deviations=_deviations(GenerationMode.RETRIEVED),
        image_quality=_image_quality(),
        manufacturability=_manufacturability(),
        rank=_ranked_result(),
        rank_explanation="coverage_pct=100%, ranked by weighted score",
    )
    return ScoredCandidate(generated=generated, scorecard=scorecard)


def _target_converged_candidate(*, optimized_zmx_resolvable: bool = False) -> ScoredCandidate:
    """A Mode3 candidate. By default its payload's `source_zmx` points at a
    filename that does NOT exist under `ZMX_AMMO_DIR` — representative of
    the real pipeline (the optimized ZMX lived in a per-job temp directory
    that's already been cleaned up by the time an export request arrives).
    `optimized_zmx_resolvable=True` exercises the (currently unreachable in
    production, but forward-compatible) path where the file does resolve."""
    assert _CASE.metadata is not None
    candidate_id = f"{_CASE.metadata.case_id}::target-converged-both"
    payload = _CASE if optimized_zmx_resolvable else _CASE.model_copy(
        update={
            "metadata": _CASE.metadata.model_copy(
                update={"source_zmx": "atelier-c1-mode3-tmp-optimized-both.zmx"}
            )
        }
    )
    codev_post_aut = {
        "post_aut.efl_y_mm": 3.79,
        "autovig.edge_used": 0.4,
        "aut_converged": "1",
    }
    accepted = FnumAcceptedFinalEvidence(
        status="measured",
        measured_fnum=1.9,
        fno_param_achieved=True,
        aut_converged=True,
        ray_traceable=True,
        effective_edge_used=0.4,
        ray_grid={
            "category": "ok", "refl_count": 0, "miss_count": 0,
            "ray_aiming_warning": False, "aperture_conflict_matched": None,
            "excerpt": None, "note": "positive measured listing evidence",
            "normal_completion": True, "abnormal_completion_matched": None,
        },
        quality_note="measured on accepted ray-retry pupil",
        optimized_zmx_path="accepted-final.zmx",
    )
    evidence = FnumLadderEvidence(
        schema="atelier-p15-fno-ladder-v1",
        target_achieved=True,
        accepted_final=accepted,
        target_efl_mm=3.8,
        fnum_target=1.9,
        stage="B",
        rung_count=3,
        fnum_tolerance_pct=8.0,
        vig_ladder=(0.0, 0.2, 0.3),
        ray_retry_vig_ladder=(0.2, 0.3, 0.4, 0.5),
        num_fields=3,
        extra_dof="both",
    )
    generated = GeneratedCandidate(
        candidate_id=candidate_id,
        mode=GenerationMode.TARGET_CONVERGED,
        source_case_id=_CASE.metadata.case_id,
        payload=payload,
        optical_extras=OpticalExtras(ri_by_field=None, codev_post_aut=codev_post_aut),
        generation_notes=["Mode3：③ target 优化标准入口"],
        fnum_ladder_evidence=evidence,
    )
    scorecard = ScorecardRow(
        candidate_id=candidate_id,
        mode=GenerationMode.TARGET_CONVERGED,
        target_deviations=_deviations(GenerationMode.TARGET_CONVERGED),
        image_quality=_image_quality(),
        manufacturability=_manufacturability(),
        rank=_ranked_result(),
        rank_explanation="coverage_pct=100%, ranked by weighted score",
    )
    return ScoredCandidate(generated=generated, scorecard=scorecard)


def _candidate_set(*candidates: ScoredCandidate) -> CandidateSet:
    mode_counts: dict[GenerationMode, int] = {}
    for sc in candidates:
        mode_counts[sc.mode] = mode_counts.get(sc.mode, 0) + 1
    summary = CandidateSetSummary(
        candidate_count=len(candidates),
        mode_counts=mode_counts,
        ranked_count=len(candidates),
        withheld_count=0,
        ri_missing_count=sum(
            1
            for sc in candidates
            if sc.scorecard.image_quality.relative_illumination.status == "unavailable"
        ),
        notes=["fixture: export test batch"],
    )
    return CandidateSet(target=_target_spec(), candidates=list(candidates), summary=summary)


def _stageb_negative_candidate() -> ScoredCandidate:
    positive = _target_converged_candidate()
    evidence = positive.generated.fnum_ladder_evidence
    assert evidence is not None
    negative_evidence = evidence.model_copy(
        update={"target_achieved": False, "accepted_final": None}
    )
    generated = positive.generated.model_copy(
        update={"fnum_ladder_evidence": negative_evidence}
    )
    deviations = [
        dev.model_copy(update={"converged_toward_target": False})
        if dev.field == "fnum"
        else dev
        for dev in positive.scorecard.target_deviations
    ]
    scorecard = positive.scorecard.model_copy(update={"target_deviations": deviations})
    return ScoredCandidate(generated=generated, scorecard=scorecard)


def _stagec_target_spec() -> TargetSpec:
    return _target_spec().model_copy(
        update={
            "image_height_mm": 3.0,
            "fov_deg": 2 * math.degrees(math.atan(3.0 / 3.8)),
        }
    )


def _stagec_offline_candidate(tmp_path: Path) -> ScoredCandidate:
    sc = _target_converged_candidate()
    source = tmp_path / "seed.zmx"
    fractions = tuple(i / 11 for i in range(12))
    zero = " ".join("0" for _ in fractions)
    source.write_text(
        "\n".join(
            (
                "FTYP 0 0 12 5 0 0 0 12",
                f"XFLN {zero}",
                "YFLN " + " ".join(str(value * 40) for value in fractions),
                f"VDXN {zero}",
                f"VDYN {zero}",
                f"VCXN {zero}",
                f"VCYN {zero}",
                "SURF 0",
            )
        )
        + "\n",
        encoding="ascii",
        newline="\n",
    )
    artifact = tmp_path / "temporary-stagec.zmx"
    target_imh = 3.0
    efl = 3.8
    derived_fov = 2 * math.degrees(math.atan(target_imh / efl))
    resolved = resolve_field_target(
        efl_mm=efl, image_height_mm=target_imh, full_fov_deg=None
    )
    reconstruction = reconstruct_image_fields(
        source_zmx=source,
        output_zmx=artifact,
        resolved_target=resolved,
    )
    evidence = StageCFieldEvidence(
        reconstruction_status="constructed",
        imh_source="constructed",
        fov_source="derived",
        efl_constraint_status="unverified",
        ray_metrics_status="pending",
        real_chief_ray_status="pending",
        rsi_status="pending",
        target_image_height_mm=3.0,
        target_efl_mm=efl,
        nominal_image_height_mm=3.0,
        derived_full_fov_deg=derived_fov,
        measured_full_fov_deg=None,
        reconstruction_applied=True,
        imh_field_valid=False,
        efl_constraint_held=False,
        ray_metrics_valid=False,
        note="FOV derived-only; [EXPERT] remains blank",
    )
    assert sc.generated.payload.metadata is not None
    payload = sc.generated.payload.model_copy(
        update={
            "metadata": sc.generated.payload.metadata.model_copy(
                update={
                    "source_zmx": artifact.name,
                    "image_height_mm": target_imh,
                    "image_height_source": "constructed",
                    "fov_deg": derived_fov,
                    "fov_source": "derived",
                    "nominal_efl_mm": efl,
                }
            )
        }
    )
    generated = GeneratedCandidate.model_validate(
        {
            **sc.generated.model_dump(),
            "payload": payload.model_dump(),
            "optimized_zmx_path": str(artifact),
            "stagec_field_reconstruction": reconstruction.model_dump(),
            "stagec_field_evidence": evidence.model_dump(),
        }
    )
    return ScoredCandidate(generated=generated, scorecard=sc.scorecard)


def _export_machine_readback(reconstruction):
    source_bytes = Path(reconstruction.source_path).read_bytes()
    reconstructed_bytes = Path(reconstruction.output_path).read_bytes()
    run_id = "stagec-export-fixture"
    config = {
        "field_type": "RIH",
        "field_count": len(reconstruction.normalized_fractions),
        "expected_samples_per_metric": 8,
        "vignetting_mode": "zero-only",
    }
    canonical = lambda value: json.dumps(  # noqa: E731
        value, sort_keys=True, separators=(",", ":")
    ).encode()
    config_fingerprint = hashlib.sha256(canonical(config)).hexdigest()
    source_sha = hashlib.sha256(source_bytes).hexdigest()
    reconstructed_sha = hashlib.sha256(reconstructed_bytes).hexdigest()
    sequence_bytes = (
        "ATELIER_STAGEC_SEQUENCE_V1\n"
        f"RUN_ID\t{run_id}\n"
        f"SOURCE_ZMX_SHA256\t{source_sha}\n"
        f"RECONSTRUCTED_ZMX_SHA256\t{reconstructed_sha}\n"
        f"CONFIG_FINGERPRINT\t{config_fingerprint}\n"
        "! synthetic export fixture body; never executable\n"
    ).encode()
    manifest = {
        "schema_id": "atelier-stagec-machine-manifest-v1",
        "run_id": run_id,
        "source_zmx_sha256": source_sha,
        "reconstructed_zmx_sha256": reconstructed_sha,
        "sequence_sha256": hashlib.sha256(sequence_bytes).hexdigest(),
        "config": config,
        "config_fingerprint": config_fingerprint,
    }
    manifest_bytes = canonical(manifest)
    listing = [
        f"ATELIER_STAGEC_RUN_BEGIN\t{run_id}",
        "SCHEMA\tatelier-stagec-listing-v1",
        f"SOURCE_ZMX_SHA256\t{manifest['source_zmx_sha256']}",
        f"RECONSTRUCTED_ZMX_SHA256\t{manifest['reconstructed_zmx_sha256']}",
        f"CONFIG_FINGERPRINT\t{config_fingerprint}",
        "TYP_FLD\tRIH",
    ]
    for index in range(config["field_count"]):
        listing += [f"FIELD_BEGIN\t{index}", f"FIELD_OK\t{index}", f"FIELD_END\t{index}"]
    listing.append(f"ATELIER_STAGEC_RUN_END\t{run_id}")
    listing_bytes = ("\n".join(listing) + "\n").encode()
    meta = {
        "schema_id": "atelier-stagec-machine-metrics-v1",
        "run_id": run_id,
        "source_zmx_sha256": manifest["source_zmx_sha256"],
        "reconstructed_zmx_sha256": manifest["reconstructed_zmx_sha256"],
        "config_fingerprint": config_fingerprint,
        "field_type": "RIH",
        "field_count": str(config["field_count"]),
        "expected_samples_per_metric": "8",
        "measured_efl_mm": str(reconstruction.target_efl_mm),
    }
    columns = (
        "record\tfield_index\tnormalized_fraction\tfield_type\tdefinition_x_ri_mm\t"
        "definition_y_ri_mm\trsi_actual_x_mm\trsi_actual_y_mm\trsi_direction_l\t"
        "rsi_direction_m\trsi_direction_n\trayrsi_return_code\trer\tbls\t"
        "rms_spot_radius_um\trms_wfe_waves\trsi_valid\trsi_attempted\tchief_valid\t"
        "chief_attempted\tspot_valid\tspot_attempted\twfe_valid\twfe_attempted\t"
        "vuy\tvly\tvux\tvlx"
    )
    metrics = [f"META\t{key}\t{value}" for key, value in meta.items()]
    metrics.append(columns)
    for index, fraction in enumerate(reconstruction.normalized_fractions):
        y = fraction * reconstruction.target_image_height_mm
        metrics.append(
            f"FIELD\t{index}\t{fraction}\tRIH\t0\t{y}\t0\t{y}\t0\t0\t1\t0\t0\t0\t"
            f"{2 + index}\t{0.2 + index / 100}\t8\t8\t8\t8\t8\t8\t8\t8\t0\t0\t0\t0"
        )
    return build_stagec_machine_readback(
        listing_bytes=listing_bytes,
        metrics_bytes=("\n".join(metrics) + "\n").encode(),
        source_zmx_bytes=source_bytes,
        reconstructed_zmx_bytes=reconstructed_bytes,
        sequence_bytes=sequence_bytes,
        manifest_bytes=manifest_bytes,
    )


def _stagec_machine_candidate(tmp_path: Path) -> ScoredCandidate:
    offline = _stagec_offline_candidate(tmp_path)
    reconstruction = offline.generated.stagec_field_reconstruction
    assert reconstruction is not None and reconstruction.output_sha256 is not None
    readback = _export_machine_readback(reconstruction)
    evidence = build_stagec_machine_evidence(
        reconstruction=reconstruction,
        readback=readback,
    )
    raw = offline.generated.model_dump()
    raw["stagec_field_evidence"] = evidence
    generated = GeneratedCandidate.model_validate(raw)
    return ScoredCandidate(generated=generated, scorecard=offline.scorecard)


# ---------------------------------------------------------------------------
# ① xlsx workbook
# ---------------------------------------------------------------------------


def test_workbook_has_summary_and_candidates_sheets_with_same_source_values():
    candidate_set = _candidate_set(_retrieved_candidate(), _target_converged_candidate())
    workbook_bytes = build_candidate_set_workbook(
        candidate_set, job_id="job-abc123", requirement="wide phone camera"
    )
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))

    assert wb.sheetnames == ["Summary", "Candidates"]

    summary_rows = [tuple(row) for row in wb["Summary"].iter_rows(values_only=True)]
    summary_text = "\n".join(str(v) for row in summary_rows for v in row if v is not None)
    assert "job-abc123" in summary_text
    assert "wide phone camera" in summary_text
    assert candidate_set.honesty_banner is None  # both modes present -> no banner in this fixture
    assert "3.8" in summary_text  # target EFL echoed

    candidates_rows = list(wb["Candidates"].iter_rows(values_only=True))
    header = candidates_rows[0]
    assert "candidate_id" in header
    assert "rank_score" in header
    body = candidates_rows[1:]
    assert len(body) == 2
    candidate_id_idx = header.index("candidate_id")
    rank_score_idx = header.index("rank_score")
    ids = {row[candidate_id_idx] for row in body}
    assert ids == {sc.scorecard.candidate_id for sc in candidate_set.candidates}
    # P17 对抗审 M3：xlsx 单元格 = 页面同一格式化器产出的同一字符串
    # （页面 rank 行显示 "score=0.812"），不再是原始 float。
    for row in body:
        assert row[rank_score_idx] == "0.812"

    # P17 sub-item 3: repeatability columns present, honest default state
    # (these fixtures never supply repeat samples -> unavailable/run_count=1,
    # same as the page's own default rendering).
    rc_idx = header.index("repeatability_run_count")
    rs_idx = header.index("repeatability_status")
    for row in body:
        assert row[rc_idx] == 1
        assert row[rs_idx] == "unavailable"


def test_workbook_nonzero_tiny_value_never_renders_as_zero():
    """P17 对抗审 M3 假零钉死（xlsx 侧）：`0 < RI < 0.0005` 曾以原始 float
    进 xlsx、页面显示 "0.000"——统一格式化器后，两侧都必须显示 "<0.001"，
    绝不显示假零。"""
    sc = _retrieved_candidate()
    tiny_ri = MetricValue(value=0.0004, status="available")
    sc = sc.model_copy(
        update={
            "scorecard": sc.scorecard.model_copy(
                update={
                    "image_quality": sc.scorecard.image_quality.model_copy(
                        update={"relative_illumination": tiny_ri}
                    )
                }
            )
        }
    )
    candidate_set = _candidate_set(sc)
    workbook_bytes = build_candidate_set_workbook(candidate_set, job_id="job-tz", requirement=None)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    rows = list(wb["Candidates"].iter_rows(values_only=True))
    header, body = rows[0], rows[1]
    ri_idx = header.index("Relative illumination (worst field)")
    assert body[ri_idx] == "<0.001"
    assert body[ri_idx] != "0.000"


def test_workbook_never_contains_pass_fail_verdict_wording():
    candidate_set = _candidate_set(_retrieved_candidate(), _target_converged_candidate())
    workbook_bytes = build_candidate_set_workbook(candidate_set, job_id="job-1", requirement=None)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))

    all_text = "\n".join(
        str(cell)
        for ws in wb.worksheets
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    assert "合格" not in all_text
    assert "良品" not in all_text
    assert "pass" not in all_text.lower()
    assert "fail" not in all_text.lower()


def test_workbook_summary_sheet_echoes_honesty_banner_when_present():
    candidate_set = _candidate_set(_retrieved_candidate())  # Mode1-only -> banner set
    assert candidate_set.honesty_banner is not None
    workbook_bytes = build_candidate_set_workbook(candidate_set, job_id="job-2", requirement=None)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    summary_text = "\n".join(
        str(v)
        for row in wb["Summary"].iter_rows(values_only=True)
        for v in row
        if v is not None
    )
    assert candidate_set.honesty_banner in summary_text


def test_workbook_handles_zero_candidates():
    empty_set = _candidate_set()
    workbook_bytes = build_candidate_set_workbook(empty_set, job_id="job-empty", requirement=None)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    candidates_rows = list(wb["Candidates"].iter_rows(values_only=True))
    assert len(candidates_rows) == 1  # header only


# ---------------------------------------------------------------------------
# ② per-candidate ZMX + reproduction .seq + README bundle
# ---------------------------------------------------------------------------


def test_bundle_zip_includes_real_zmx_for_retrieved_candidate():
    sc = _retrieved_candidate()
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "candidate.zmx" in names
        assert "README.txt" in names
        assert "reproduction.seq" not in names  # Mode1: nothing to reproduce

        assert _CASE.metadata is not None
        real_bytes = (ZMX_AMMO_DIR / _CASE.metadata.source_zmx).read_bytes()
        assert zf.read("candidate.zmx") == real_bytes

        readme = zf.read("README.txt").decode("utf-8")
    assert "candidate.zmx: included" in readme
    assert "not applicable" in readme  # .seq n/a for Mode1
    assert "合格" not in readme
    assert "良品" not in readme
    assert "pass" not in readme.lower()
    assert "fail" not in readme.lower()


def test_bundle_zip_omits_zmx_and_explains_when_not_persisted():
    """The known Mode3 limitation: the optimized ZMX lived in a per-job temp
    directory already cleaned up by export time — the bundle must fail
    closed (no candidate.zmx entry) with an honest README, never a broken or
    silently-missing artifact."""
    sc = _target_converged_candidate(optimized_zmx_resolvable=False)
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "candidate.zmx" not in names
        assert "README.txt" in names
        readme = zf.read("README.txt").decode("utf-8")
    assert "NOT included" in readme
    assert "persistence failure" in readme


def test_bundle_zip_includes_zmx_once_mode3_artifact_is_persisted():
    """Forward-compatibility check: if a candidate's payload does resolve to
    a real file under ZMX_AMMO_DIR (e.g. once the generators.py-side
    persistence gap is closed), the same mode-agnostic resolution path picks
    it up with no code change needed."""
    sc = _target_converged_candidate(optimized_zmx_resolvable=True)
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "candidate.zmx" in zf.namelist()


def test_bundle_zip_reads_persisted_mode3_artifact_path(tmp_path: Path):
    expected = b"persistent optimized zmx bytes\r\n"
    artifact = tmp_path / "job" / "candidates" / "candidate-key" / "candidate.zmx"
    artifact.parent.mkdir(parents=True)
    artifact.write_bytes(expected)
    sc = _target_converged_candidate().model_copy(
        update={
            "generated": _target_converged_candidate().generated.model_copy(
                update={"optimized_zmx_path": str(artifact)}
            )
        }
    )
    with zipfile.ZipFile(io.BytesIO(build_candidate_bundle_zip(sc, target=_target_spec()))) as zf:
        assert zf.read("candidate.zmx") == expected


def test_bundle_zip_reconstructs_mode3_reproduction_seq_when_provenance_complete():
    sc = _target_converged_candidate()
    target = _target_spec()
    zip_bytes = build_candidate_bundle_zip(sc, target=target)

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "reproduction.seq" in names
        seq_text = zf.read("reproduction.seq").decode("ascii")
        readme = zf.read("README.txt").decode("utf-8")

    assert seq_text.strip()  # non-empty macro text
    assert f"{target.efl_mm}" in seq_text or f"{target.efl_mm:.6f}" in seq_text
    assert "FNO 1.9" in seq_text
    assert "VUY" in seq_text and "VLY" in seq_text
    assert "reproduction.seq: included" in readme
    assert "accepted_final.measured_fnum: 1.900" in readme
    assert "accepted_final.effective_edge_used: 0.400" in readme
    assert '"category": "ok"' in readme
    assert "measured on accepted ray-retry pupil" in readme


def test_stageb_negative_evidence_exports_false_and_omits_reproduction_seq():
    sc = _stageb_negative_candidate()
    workbook = openpyxl.load_workbook(
        io.BytesIO(
            build_candidate_set_workbook(
                _candidate_set(sc), job_id="stageb-negative", requirement=None
            )
        )
    )
    rows = list(workbook["Candidates"].iter_rows(values_only=True))
    header, values = rows[0], rows[1]
    assert values[header.index("fnum_ladder_target_achieved")] is False
    assert values[header.index("fnum_accepted_measured_fnum")] == "N/A"
    assert values[header.index("fnum_accepted_effective_edge_used")] == "N/A"
    assert values[header.index("fnum_accepted_ray_grid")] == "N/A"
    assert values[header.index("fnum_accepted_quality_note")] == "N/A"

    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(sc, target=_target_spec()))
    ) as zf:
        assert "reproduction.seq" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "fnum_ladder.target_achieved: False" in readme
    assert "accepted_final.measured_fnum: N/A" in readme
    assert "did not produce an accepted_final rung" in readme


def test_seed_zmx_resolution_matches_disk_name_exactly():
    """CI 红修根因钉死（2026-07-11，run 29130312114）：`data/zmx/` 扩展名
    混合大小写（5 颗 `.ZMX` / 437 颗 `.zmx`），本 fixture 的 seed 恰是
    `.ZMX` 之一。旧实现从 `case_id` 合成 `f"{case_id}.zmx"`——Windows 大小写
    不敏感文件系统上 `is_file()` 照样为真（本地绿），Ubuntu CI 大小写敏感
    直接 miss（CI 红）。本测试用**精确目录清单比对**（大小写敏感，两个平台
    行为一致），Windows 上也能抓住任何重新引入的合成文件名。"""
    sc = _target_converged_candidate()
    path = _resolve_seed_zmx_path(sc)
    assert path is not None, "seed ZMX must resolve via the library's exact source_zmx"
    assert path.name in os.listdir(ZMX_AMMO_DIR), (
        f"resolved name {path.name!r} must exact-match a directory entry "
        "(case-sensitive comparison — Windows' case-insensitive is_file() must "
        "not mask a wrong-case synthesized filename)"
    )
    # And it must be the seed's own recorded filename, not a guess.
    assert sc.generated.source_case_id is not None
    seed_case = next(
        c
        for c in load_case_library()
        if c.metadata is not None and c.metadata.case_id == sc.generated.source_case_id
    )
    assert seed_case.metadata is not None
    assert path.name == seed_case.metadata.source_zmx


def test_bundle_zip_seq_fails_closed_without_stageb_evidence():
    """A legacy post-AUT snapshot cannot substitute for closed Stage B evidence."""
    sc = _target_converged_candidate()
    sc = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(
                update={
                    "optical_extras": OpticalExtras(
                        ri_by_field=None, codev_post_aut={"post_aut.efl_y_mm": 3.79}
                    ),
                    "fnum_ladder_evidence": None,
                }
            )
        }
    )
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "reproduction.seq" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "NOT included" in readme
    assert "validated Stage B FNO-ladder evidence missing" in readme


def test_stagec_web_xlsx_bundle_sources_are_honest_and_replay_fails_closed(
    tmp_path: Path,
):
    sc = _stagec_offline_candidate(tmp_path)
    workbook = openpyxl.load_workbook(
        io.BytesIO(
            build_candidate_set_workbook(
                _candidate_set(sc), job_id="stagec-offline", requirement=None
            )
        ),
        read_only=True,
        data_only=True,
    )
    rows = list(workbook["Candidates"].iter_rows(values_only=True))
    header, values = rows[0], rows[1]
    assert values[header.index("stagec_machine_execution_status")] == "blocked"
    assert values[header.index("stagec_reconstruction_status")] == "constructed"
    assert values[header.index("stagec_imh_source")] == "constructed"
    assert values[header.index("stagec_imh_achieved")] is False
    assert values[header.index("stagec_target_efl_mm")] == "3.800"
    assert values[header.index("stagec_fov_source")] == "derived"
    assert values[header.index("stagec_fov_deg")] == f"{sc.generated.stagec_field_evidence.derived_full_fov_deg:.3f}"
    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None
    assert sc.generated.stagec_field_evidence.derived_full_fov_deg == pytest.approx(
        2
        * math.degrees(
            math.atan(
                reconstruction.target_image_height_mm / reconstruction.target_efl_mm
            )
        )
    )
    assert not math.isclose(
        reconstruction.target_efl_mm,
        sc.generated.payload.paraxial.effective_focal_length_mm,
        rel_tol=1e-12,
    ), "fixture must prove post-run payload EFL is not the FOV provenance source"
    assert values[header.index("stagec_real_chief_ray_status")] == "pending"
    assert values[header.index("stagec_rsi_status")] == "pending"

    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(sc, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" in zf.namelist()
        assert "reproduction.seq" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "FOV: derived/measured only; never optimized/converged" in readme
    assert "Stage C CODE V field syntax" in readme
    assert "[EXPERT]" in readme


def test_stagec_machine_evidence_crosses_candidate_and_export_boundaries(
    tmp_path: Path,
) -> None:
    sc = _stagec_machine_candidate(tmp_path)
    evidence = sc.generated.stagec_field_evidence
    assert evidence is not None and evidence.evidence_kind == "machine"
    assert evidence.image_height_achieved is False
    assert evidence.machine_execution_status == "parsed-unverified"
    persisted = sc.generated.model_dump(mode="json")
    persisted["stagec_field_evidence"]["imh_field_valid"] = False
    restored = GeneratedCandidate.model_validate(persisted)
    assert restored.stagec_field_evidence is not None
    assert restored.stagec_field_evidence.image_height_achieved is False
    deviations = {item.field: item for item in sc.scorecard.target_deviations}
    assert deviations["imh"].converged_toward_target is False
    assert deviations["fov"].converged_toward_target is False
    workbook = openpyxl.load_workbook(
        io.BytesIO(
            build_candidate_set_workbook(
                _candidate_set(sc), job_id="stagec-machine", requirement=None
            )
        ),
        read_only=True,
        data_only=True,
    )
    rows = list(workbook["Candidates"].iter_rows(values_only=True))
    header, values = rows[0], rows[1]
    assert values[header.index("stagec_machine_execution_status")] == "parsed-unverified"
    assert values[header.index("stagec_imh_source")] == "constructed-unverified"
    assert values[header.index("stagec_imh_achieved")] is False
    assert values[header.index("stagec_fov_source")] == "derived"
    machine_status_columns = (
        "stagec_machine_execution_status",
        "stagec_imh_source",
        "stagec_real_chief_ray_status",
        "stagec_rsi_status",
    )
    for column in machine_status_columns:
        assert values[header.index(column)] not in {
            "verified",
            "constructed-machine-verified",
            "zero-verified",
            "nonzero-verified",
        }

    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(sc, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" in zf.namelist()
        assert "reproduction.seq" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert '"config_fingerprint": "' in readme
    assert '"ray_classification": "valid"' in readme
    assert '"machine_execution_status": "parsed-unverified"' in readme
    assert '"machine_execution_status": "verified"' not in readme
    assert '"classification": "zero-parsed-unverified"' in readme
    for forbidden in (
        '"classification": "zero-verified"',
        '"classification": "nonzero-verified"',
        '"imh_source": "constructed-machine-verified"',
        '"real_chief_ray_status": "verified"',
        '"rsi_status": "verified"',
        '"ray_metrics_status": "verified"',
    ):
        assert forbidden not in readme
    assert "FOV: derived/measured only; never optimized/converged" in readme
    assert "[EXPERT]" in readme


def test_export_entry_revalidates_model_copy_forged_mode_and_emits_no_verified_output(
    tmp_path: Path,
) -> None:
    sc = _stagec_machine_candidate(tmp_path)
    forged = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(update={"mode": GenerationMode.RETRIEVED})
        }
    )
    forged_set = _candidate_set(sc).model_copy(update={"candidates": [forged]})
    with pytest.raises(ValueError):
        build_candidate_set_workbook(
            forged_set, job_id="forged-mode", requirement=None
        )
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(forged, target=_stagec_target_spec()))
    ) as zf:
        assert zf.namelist() == ["README.txt"]
        readme = zf.read("README.txt").decode("utf-8")
    assert "STRICT VALIDATION REJECTION" in readme
    assert "candidate.zmx: NOT included" in readme
    assert '"verified"' not in readme


def test_export_entry_revalidates_model_copy_forged_stagec_contradiction(
    tmp_path: Path,
) -> None:
    sc = _stagec_offline_candidate(tmp_path)
    evidence = sc.generated.stagec_field_evidence
    assert isinstance(evidence, StageCFieldEvidence)
    forged_evidence = evidence.model_copy(
        update={
            "target_image_height_mm": 3.1,
            "nominal_image_height_mm": 3.1,
            "derived_full_fov_deg": 2 * math.degrees(math.atan(3.1 / 3.8)),
        }
    )
    forged = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(
                update={"stagec_field_evidence": forged_evidence}
            )
        }
    )
    with pytest.raises(ValueError, match="target differs"):
        build_candidate_set_workbook(
            _candidate_set(forged), job_id="forged-stagec", requirement=None
        )
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(forged, target=_stagec_target_spec()))
    ) as zf:
        assert zf.namelist() == ["README.txt"]
        readme = zf.read("README.txt").decode("utf-8")
    assert "STRICT VALIDATION REJECTION" in readme
    assert "candidate.zmx: NOT included" in readme


def test_stagec_bundle_withholds_candidate_zmx_after_artifact_tamper(tmp_path: Path):
    sc = _stagec_offline_candidate(tmp_path)
    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None and reconstruction.output_path is not None
    Path(reconstruction.output_path).write_bytes(b"tampered-after-validation")
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(sc, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "source/output hash" in readme
    assert "candidate.zmx withheld" in readme


def test_stagec_bundle_withholds_candidate_zmx_after_source_tamper(tmp_path: Path):
    sc = _stagec_offline_candidate(tmp_path)
    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None
    Path(reconstruction.source_path).write_bytes(b"tampered-source")
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(sc, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "candidate.zmx withheld" in readme


def test_stagec_bundle_rejects_arbitrary_bytes_even_with_self_reported_hash(tmp_path: Path):
    sc = _stagec_offline_candidate(tmp_path)
    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None and reconstruction.output_path is not None
    artifact = Path(reconstruction.output_path)
    artifact.write_bytes(b"arbitrary-but-self-hashed\n")
    forged_reconstruction = reconstruction.model_copy(
        update={"output_sha256": hashlib.sha256(artifact.read_bytes()).hexdigest()}
    )
    forged_generated = sc.generated.model_copy(
        update={"stagec_field_reconstruction": forged_reconstruction}
    )
    forged = sc.model_copy(update={"generated": forged_generated})
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(forged, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "candidate.zmx withheld" in readme


@pytest.mark.parametrize(
    "forged_ftyp",
    ["FTYP 3 0 999", "FTYP 3.5 0 12"],
)
def test_stagec_candidate_and_export_reject_self_hashed_ftyp_forgery(
    tmp_path: Path, forged_ftyp: str,
):
    sc = _stagec_offline_candidate(tmp_path)
    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None and reconstruction.output_path is not None
    artifact = Path(reconstruction.output_path)
    artifact.write_text(
        artifact.read_text(encoding="ascii").replace("FTYP 3 0 12", forged_ftyp),
        encoding="ascii",
        newline="\n",
    )
    forged_reconstruction = reconstruction.model_copy(
        update={"output_sha256": hashlib.sha256(artifact.read_bytes()).hexdigest()}
    )
    raw = sc.generated.model_dump()
    raw["stagec_field_reconstruction"] = forged_reconstruction.model_dump()
    with pytest.raises(ValueError, match="artifact bytes are invalid"):
        GeneratedCandidate.model_validate(raw)

    forged_generated = sc.generated.model_copy(
        update={"stagec_field_reconstruction": forged_reconstruction}
    )
    forged = sc.model_copy(update={"generated": forged_generated})
    with zipfile.ZipFile(
        io.BytesIO(build_candidate_bundle_zip(forged, target=_stagec_target_spec()))
    ) as zf:
        assert "candidate.zmx" not in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "candidate.zmx withheld" in readme


def test_stagec_candidate_rejects_target_profile_and_artifact_path_mismatch(tmp_path: Path):
    sc = _stagec_offline_candidate(tmp_path)
    raw = sc.generated.model_dump()
    raw["stagec_field_evidence"]["target_image_height_mm"] = 3.1
    raw["stagec_field_evidence"]["nominal_image_height_mm"] = 3.1
    raw["stagec_field_evidence"]["derived_full_fov_deg"] = 2 * math.degrees(
        math.atan(3.1 / 3.8)
    )
    with pytest.raises(ValueError, match="target differs"):
        GeneratedCandidate.model_validate(raw)

    raw = sc.generated.model_dump()
    raw["stagec_field_evidence"]["target_efl_mm"] = 3.9
    raw["stagec_field_evidence"]["derived_full_fov_deg"] = 2 * math.degrees(
        math.atan(3.0 / 3.9)
    )
    with pytest.raises(ValueError, match="target EFL differs"):
        GeneratedCandidate.model_validate(raw)

    raw = sc.generated.model_dump()
    raw["stagec_field_reconstruction"]["num_fields"] = 11
    with pytest.raises(ValueError, match="profile length"):
        GeneratedCandidate.model_validate(raw)

    other = tmp_path / "other.zmx"
    other.write_bytes(Path(sc.generated.optimized_zmx_path).read_bytes())
    raw = sc.generated.model_dump()
    raw["optimized_zmx_path"] = str(other)
    with pytest.raises(ValueError, match="must be the Stage C reconstruction output"):
        GeneratedCandidate.model_validate(raw)

    reconstruction = sc.generated.stagec_field_reconstruction
    assert reconstruction is not None and reconstruction.output_path is not None
    artifact = Path(reconstruction.output_path)
    artifact.write_bytes(b"self-hashed-but-not-a-zmx\n")
    raw = sc.generated.model_dump()
    raw["stagec_field_reconstruction"]["output_sha256"] = hashlib.sha256(
        artifact.read_bytes()
    ).hexdigest()
    with pytest.raises(ValueError, match="artifact bytes are invalid"):
        GeneratedCandidate.model_validate(raw)


def test_bundle_zip_seq_fails_closed_when_edge_used_nonfinite():
    """NaN/inf edge_used 同属"不可重建"（fail closed），不是可用的裁瞳量。"""
    sc = _target_converged_candidate()
    evidence = sc.generated.fnum_ladder_evidence
    assert evidence is not None and evidence.accepted_final is not None
    invalid_evidence = evidence.model_copy(
        update={
            "accepted_final": evidence.accepted_final.model_copy(
                update={"effective_edge_used": float("nan")}
            )
        }
    )
    sc = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(
                update={
                    "optical_extras": OpticalExtras(
                        ri_by_field=None,
                        codev_post_aut={"autovig.edge_used": float("nan")},
                    ),
                    "fnum_ladder_evidence": invalid_evidence,
                }
            )
        }
    )
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "reproduction.seq" not in zf.namelist()


def test_bundle_zip_seq_included_when_edge_used_is_zero():
    """`edge_used == 0.0` 是合法完整 provenance（实际跑次零裁瞳），必须与
    "缺失"严格区分——照常交付 .seq（`_autovig_profile(0)` → 无渐晕 = 忠实
    复现该跑次）。"""
    sc = _target_converged_candidate()
    evidence = sc.generated.fnum_ladder_evidence
    assert evidence is not None and evidence.accepted_final is not None
    zero_evidence = evidence.model_copy(
        update={
            "accepted_final": evidence.accepted_final.model_copy(
                update={"effective_edge_used": 0.0}
            )
        }
    )
    sc = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(
                update={
                    "optical_extras": OpticalExtras(
                        ri_by_field=None,
                        codev_post_aut={"autovig.edge_used": 0.0},
                    ),
                    "fnum_ladder_evidence": zero_evidence,
                }
            )
        }
    )
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "reproduction.seq" in zf.namelist()
        readme = zf.read("README.txt").decode("utf-8")
    assert "reproduction.seq: included" in readme


def test_bundle_zip_seq_omitted_when_seed_zmx_unresolvable():
    sc = _target_converged_candidate()
    sc = sc.model_copy(
        update={
            "generated": sc.generated.model_copy(update={"source_case_id": "no-such-seed-case"})
        }
    )
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "reproduction.seq" not in names
        readme = zf.read("README.txt").decode("utf-8")
    assert "NOT included" in readme
    assert "seed ZMX not resolvable" in readme  # concrete missing-provenance reason (M2)


def test_bundle_zip_seq_not_applicable_for_retrieved_candidate():
    sc = _retrieved_candidate()
    zip_bytes = build_candidate_bundle_zip(sc, target=_target_spec())
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        assert "reproduction.seq" not in zf.namelist()
