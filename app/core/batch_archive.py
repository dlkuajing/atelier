"""P18-1: file-backed archive storage for batch/job/[EXPERT]-verdict records.

No external DB — every record is a JSON file under `settings.batch_archive_dir`
(default `var/batch-archive`, gitignored). Layout, one directory per batch:

    <root>/<batch_id>/batch.json        - BatchRecord
    <root>/<batch_id>/targets.json      - resolved target list, frozen at creation
    <root>/<batch_id>/jobs/*.json       - one BatchJobRecord per target attempt
    <root>/<batch_id>/verdicts/*.json   - one ExpertVerdict per (job_id, candidate_key)

Every write is atomic (temp file + `os.replace`, same directory so the
rename is same-filesystem on both POSIX and Windows) — a process killed at
any point leaves either the previous fully-written file or nothing, never a
truncated/corrupt one. Nothing is cached in memory: every read hits disk, so
a fresh `BatchArchive` instance after a crash reconstructs identical state
from whatever files finished writing before the kill (crash-safety
requirement, P18 batch brief 交付1).

**[EXPERT] 判定权红线**: `ExpertVerdict.verdict_text`/`reviewer` are
non-blank-required free text — this module never fabricates, defaults, or
pre-fills a verdict. "Not yet recorded" is represented by the *absence* of a
verdict file (`get_verdict` returns `None`), never by a constructible
placeholder object with an empty/zero value standing in for "unreviewed".
"""

from __future__ import annotations

import io
import json
import os
import re
from collections.abc import Mapping, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field, ValidationError, model_validator

from app.core.config import settings

BatchStatus = Literal["running", "completed", "budget_exhausted", "aborted"]
#: `degraded` (P18 对抗审 BLOCKER-1): the engine returned candidates, but at
#: least one *requested* generation mode produced nothing — e.g. a real batch
#: whose Mode3/CODE V leg silently fell away (CODE V unavailable). Such a job
#: must never book as `succeeded`; the degradation reason is mandatory.
JobStatus = Literal["queued", "running", "succeeded", "degraded", "failed"]
FailureCategory = Literal["preflight", "engine", "timeout", "exception"]

_NON_PASS_FAIL_BANNER = (
    "本表仅量化数据，不代为判定——量产可用性判断权与 [EXPERT] 背书"
    "始终在资深设计师手里（AGENTS.md 北极星条款）。"
)


class BatchArchiveError(RuntimeError):
    """Raised for archive-layer failures: unknown batch/job/verdict lookups,
    or a `create_batch` call that collides with an existing batch_id."""


class CandidateSetUnavailableError(BatchArchiveError):
    """The job's persisted `CandidateSet` cannot be loaded (no
    `candidate_set_pointer`, file missing, or corrupt payload) — verdict
    writes fail closed: a verdict cannot anchor to candidates we cannot
    enumerate (P18 对抗审 MAJOR-2)."""


class UnknownCandidateKeyError(BatchArchiveError):
    """The verdict's `candidate_key` does not match any candidate in the
    job's persisted `CandidateSet` — a ghost [EXPERT] verdict must never be
    written (P18 对抗审 MAJOR-2)."""


# ---------------------------------------------------------------------------
# Records
# ---------------------------------------------------------------------------


class BatchJobFailure(BaseModel):
    """Machine-recorded failure classification for one job — a diagnostic
    fact, never a judgment call. `category` values map onto the batch
    runner's honest-ledger contract (P18-2): `preflight` (target spec failed
    validation before the engine ever ran), `engine` (the engine raised, or
    produced a result the engine itself flags as a failed attempt — e.g. zero
    candidates), `timeout` (exceeded the configured per-job wall limit), or
    `exception` (unclassified — the batch runner's own safety-net catch)."""

    category: FailureCategory
    message: str


class BatchJobRecord(BaseModel):
    """One target's orchestration attempt within a batch.

    `target_spec` is the *raw* input entry (not a validated `TargetSpec`) so
    a `status="failed"`/`failure.category="preflight"` record — which by
    definition never reached a valid `TargetSpec` — can still show what was
    attempted. `candidate_set_pointer` is a path to the full `CandidateSet`
    JSON persisted alongside the job's artifacts, kept out of this
    lightweight record so listing a batch's jobs never loads every candidate
    set into memory.
    """

    job_id: str
    batch_id: str
    target_index: int = Field(..., ge=0)
    target_label: str
    target_spec: dict[str, object]
    status: JobStatus
    created_at: str
    updated_at: str
    result_summary: dict[str, object] | None = None
    candidate_set_pointer: str | None = None
    artifact_dir: str | None = None
    attempt: int = Field(
        1,
        ge=1,
        description=(
            "本 job 的第几次尝试（MAJOR-3 attempt provenance）：每次尝试写进自己的 "
            "attempt-N 工件子目录，timeout 后旧目录绝不复用——一个未被杀死的超时"
            "线程只能继续写它自己那份已被判死的 attempt 目录，碰不到下一次尝试。"
        ),
    )
    failure: BatchJobFailure | None = None
    degradation: str | None = Field(
        None,
        description=(
            "status=degraded 时必填的人读原因（如 'requested mode target-converged "
            "produced no candidates'）；其它状态必须为 None（BLOCKER-1 诚实账本）"
        ),
    )

    @model_validator(mode="after")
    def _failure_consistent_with_status(self) -> BatchJobRecord:
        if self.status == "failed" and self.failure is None:
            raise ValueError(
                "status=failed 时 failure 必须提供分类（诚实账本，不可留空）"
            )
        if self.status != "failed" and self.failure is not None:
            raise ValueError("failure 只能在 status=failed 时设置")
        if self.status == "degraded" and not (self.degradation or "").strip():
            raise ValueError(
                "status=degraded 时 degradation 原因必填（BLOCKER-1：静默降级不可无痕）"
            )
        if self.status != "degraded" and self.degradation is not None:
            raise ValueError("degradation 只能在 status=degraded 时设置")
        return self


class BatchRecord(BaseModel):
    """One night-batch run's top-level ledger entry. `target_count` is the
    resolved target list's length at creation time — the list itself lives
    in `targets.json` alongside (see `BatchArchive.get_targets`), frozen so
    `--resume` always replays the exact same indexed target set regardless
    of what a later invocation's CLI flags say."""

    batch_id: str
    created_at: str
    updated_at: str
    target_source: str
    target_count: int = Field(..., ge=0)
    status: BatchStatus
    engine: str
    notes: list[str] = Field(default_factory=list)


class ExpertVerdict(BaseModel):
    """[EXPERT] 判定权红线的存储面：候选级自由文本 verdict，从不由 AI/系统
    预填或默认。`verdict_text`/`reviewer` 均要求非空白字符串——空字符串等价
    于"未录入"，但那个状态由**没有对应文件**表达（`BatchArchive.get_verdict`
    返回 `None`），不是一个允许被构造出来的"空 verdict"对象。没有枚举/默认
    值：verdict 是自由文本，不是预置的"合格/良品"选项。"""

    job_id: str
    candidate_key: str
    verdict_text: str = Field(..., min_length=1)
    reviewer: str = Field(..., min_length=1)
    recorded_at: str
    note: str | None = None

    @model_validator(mode="after")
    def _not_blank(self) -> ExpertVerdict:
        if not self.verdict_text.strip() or not self.reviewer.strip():
            raise ValueError(
                "verdict_text/reviewer 不可为空白字符串（[EXPERT] 红线：未录入"
                "必须显式留白，不能靠空串冒充已录入）"
            )
        return self


# ---------------------------------------------------------------------------
# Disk I/O helpers
# ---------------------------------------------------------------------------


def _utc_now_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def _new_batch_id() -> str:
    return f"{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}-{uuid4().hex[:8]}"


def _safe_component(value: str) -> str:
    """Collapse a job_id/candidate_key into a filesystem-safe path
    component — candidate ids embed `::`, and an unsanitized batch/job id
    could otherwise walk out of `root` via `..`."""
    collapsed = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    return collapsed or "_"


def _atomic_write_json(path: Path, payload: Mapping[str, object]) -> None:
    """Temp-file-then-rename write: `os.replace` is atomic on both POSIX and
    Windows when source/dest share a filesystem (guaranteed here — the temp
    file is always written next to its final name), so a process killed
    mid-write leaves either the previous complete file or nothing, never a
    half-written one. The temp name keeps the `.json` suffix out of its
    final position, so `glob("*.json")` readers never pick up an in-flight
    write."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + f".tmp-{uuid4().hex}")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, path)


def _read_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# BatchArchive — CRUD
# ---------------------------------------------------------------------------


class BatchArchive:
    """File-backed persistence for batches/jobs/[EXPERT] verdicts. Stateless
    between calls (no in-memory cache) by design — every read hits disk, so
    a fresh instance after a crash sees exactly what made it to disk."""

    def __init__(self, root: Path | None = None) -> None:
        self.root = Path(root) if root is not None else settings.batch_archive_dir

    # -- paths --

    def _batch_dir(self, batch_id: str) -> Path:
        return self.root / _safe_component(batch_id)

    def _batch_file(self, batch_id: str) -> Path:
        return self._batch_dir(batch_id) / "batch.json"

    def _targets_file(self, batch_id: str) -> Path:
        return self._batch_dir(batch_id) / "targets.json"

    def _jobs_dir(self, batch_id: str) -> Path:
        return self._batch_dir(batch_id) / "jobs"

    def _job_file(self, batch_id: str, job_id: str) -> Path:
        return self._jobs_dir(batch_id) / f"{_safe_component(job_id)}.json"

    def _verdicts_dir(self, batch_id: str) -> Path:
        return self._batch_dir(batch_id) / "verdicts"

    def _verdict_file(self, batch_id: str, job_id: str, candidate_key: str) -> Path:
        name = f"{_safe_component(job_id)}__{_safe_component(candidate_key)}.json"
        return self._verdicts_dir(batch_id) / name

    # -- batch --

    def create_batch(
        self,
        *,
        target_source: str,
        targets: Sequence[Mapping[str, object]],
        engine: str,
        batch_id: str | None = None,
    ) -> BatchRecord:
        """Create a new batch, freezing `targets` to disk (`targets.json`) so
        `--resume` never has to re-derive the target list from CLI flags."""
        resolved_id = batch_id or _new_batch_id()
        if self._batch_file(resolved_id).is_file():
            raise BatchArchiveError(f"batch already exists: {resolved_id}")
        now = _utc_now_iso()
        record = BatchRecord(
            batch_id=resolved_id,
            created_at=now,
            updated_at=now,
            target_source=target_source,
            target_count=len(targets),
            status="running",
            engine=engine,
            notes=[],
        )
        self._batch_dir(resolved_id).mkdir(parents=True, exist_ok=True)
        _atomic_write_json(self._targets_file(resolved_id), {"targets": list(targets)})
        _atomic_write_json(self._batch_file(resolved_id), record.model_dump(mode="json"))
        return record

    def get_batch(self, batch_id: str) -> BatchRecord:
        path = self._batch_file(batch_id)
        if not path.is_file():
            raise BatchArchiveError(f"batch not found: {batch_id}")
        return BatchRecord.model_validate(_read_json(path))

    def list_batches(self) -> list[BatchRecord]:
        """Newest-first. A corrupt/partial `batch.json` (should not occur
        given atomic writes, but a manually-truncated file is still
        possible) is skipped rather than raised — one bad record must not
        take down the whole list page."""
        if not self.root.is_dir():
            return []
        batches: list[BatchRecord] = []
        for child in self.root.iterdir():
            batch_file = child / "batch.json"
            if not batch_file.is_file():
                continue
            try:
                batches.append(BatchRecord.model_validate(_read_json(batch_file)))
            except (json.JSONDecodeError, ValidationError):
                continue
        batches.sort(key=lambda b: b.created_at, reverse=True)
        return batches

    def update_batch(self, batch_id: str, **changes: object) -> BatchRecord:
        record = self.get_batch(batch_id)
        updated = record.model_copy(update={**changes, "updated_at": _utc_now_iso()})
        _atomic_write_json(self._batch_file(batch_id), updated.model_dump(mode="json"))
        return updated

    def get_targets(self, batch_id: str) -> list[dict[str, object]]:
        path = self._targets_file(batch_id)
        if not path.is_file():
            raise BatchArchiveError(f"targets not found for batch: {batch_id}")
        payload = _read_json(path)
        targets = payload.get("targets", [])
        assert isinstance(targets, list)
        return list(targets)

    # -- jobs --

    def put_job(self, job: BatchJobRecord) -> None:
        """Create-or-overwrite (idempotent) — the batch runner calls this
        once to mark a job `running` and again with the terminal
        succeeded/failed snapshot; both are just atomic full-file writes."""
        _atomic_write_json(self._job_file(job.batch_id, job.job_id), job.model_dump(mode="json"))

    def get_job(self, batch_id: str, job_id: str) -> BatchJobRecord:
        path = self._job_file(batch_id, job_id)
        if not path.is_file():
            raise BatchArchiveError(f"job not found: {batch_id}/{job_id}")
        return BatchJobRecord.model_validate(_read_json(path))

    def list_jobs(self, batch_id: str) -> list[BatchJobRecord]:
        """Sorted by `target_index` (not filename) so callers get a stable,
        human-meaningful order regardless of filesystem directory order."""
        jobs_dir = self._jobs_dir(batch_id)
        if not jobs_dir.is_dir():
            return []
        jobs: list[BatchJobRecord] = []
        for path in sorted(jobs_dir.glob("*.json")):
            try:
                jobs.append(BatchJobRecord.model_validate(_read_json(path)))
            except (json.JSONDecodeError, ValidationError):
                continue
        jobs.sort(key=lambda j: j.target_index)
        return jobs

    # -- expert verdicts --

    def candidate_keys_for_job(self, batch_id: str, job_id: str) -> frozenset[str]:
        """Enumerate the candidate ids in the job's persisted `CandidateSet`
        (raw JSON traversal of `candidate_set_pointer` — cheap, no dependency
        on the orchestration models). MAJOR-2's verdict anchor: raises
        `BatchArchiveError` for an unknown job and
        `CandidateSetUnavailableError` when the pointer is absent, the file
        is gone, or the payload is malformed — a verdict target set we
        cannot enumerate means no verdict can be written (fail closed). An
        honestly *empty* candidates list is valid and returns an empty set
        (every key is then unknown)."""
        job = self.get_job(batch_id, job_id)
        if not job.candidate_set_pointer:
            raise CandidateSetUnavailableError(
                f"job has no persisted candidate set: {batch_id}/{job_id}"
            )
        path = Path(job.candidate_set_pointer)
        if not path.is_file():
            raise CandidateSetUnavailableError(
                f"candidate set file missing on disk: {job.candidate_set_pointer}"
            )
        try:
            payload = _read_json(path)
        except (json.JSONDecodeError, OSError) as exc:
            raise CandidateSetUnavailableError(
                f"candidate set file unreadable/corrupt: {job.candidate_set_pointer}"
            ) from exc
        candidates = payload.get("candidates")
        if not isinstance(candidates, list):
            raise CandidateSetUnavailableError(
                f"candidate set payload malformed (no candidates list): {job.candidate_set_pointer}"
            )
        keys: set[str] = set()
        for entry in candidates:
            candidate_id = (
                entry.get("scorecard", {}).get("candidate_id")
                if isinstance(entry, dict) and isinstance(entry.get("scorecard"), dict)
                else None
            )
            if not isinstance(candidate_id, str) or not candidate_id:
                raise CandidateSetUnavailableError(
                    "candidate set payload malformed (candidate without scorecard."
                    f"candidate_id): {job.candidate_set_pointer}"
                )
            keys.add(candidate_id)
        return frozenset(keys)

    def put_verdict(self, verdict: ExpertVerdict, *, batch_id: str) -> None:
        """Fails loud if the verdict doesn't anchor to reality (P18 对抗审
        MAJOR-2 — a ghost [EXPERT] verdict must be unwritable at the storage
        layer, not just discouraged at the web layer):

        - unknown job -> `BatchArchiveError`
        - job's candidate set absent/corrupt -> `CandidateSetUnavailableError`
        - `candidate_key` not in the persisted set -> `UnknownCandidateKeyError`
        """
        keys = self.candidate_keys_for_job(batch_id, verdict.job_id)
        if verdict.candidate_key not in keys:
            raise UnknownCandidateKeyError(
                f"candidate_key {verdict.candidate_key!r} is not a candidate of "
                f"{batch_id}/{verdict.job_id} (persisted candidate set has "
                f"{len(keys)} candidates)"
            )
        _atomic_write_json(
            self._verdict_file(batch_id, verdict.job_id, verdict.candidate_key),
            verdict.model_dump(mode="json"),
        )

    def get_verdict(self, batch_id: str, job_id: str, candidate_key: str) -> ExpertVerdict | None:
        path = self._verdict_file(batch_id, job_id, candidate_key)
        if not path.is_file():
            return None
        return ExpertVerdict.model_validate(_read_json(path))

    def list_verdicts(self, batch_id: str, *, job_id: str | None = None) -> list[ExpertVerdict]:
        verdicts_dir = self._verdicts_dir(batch_id)
        if not verdicts_dir.is_dir():
            return []
        out: list[ExpertVerdict] = []
        for path in sorted(verdicts_dir.glob("*.json")):
            try:
                verdict = ExpertVerdict.model_validate(_read_json(path))
            except (json.JSONDecodeError, ValidationError):
                continue
            if job_id is None or verdict.job_id == job_id:
                out.append(verdict)
        return out


# ---------------------------------------------------------------------------
# xlsx export — machine columns and [EXPERT] columns physically separated
# into different sheets (never mixed in the same row), per the North Star's
# expert-verdict-ownership rule. Absent verdicts render as blank rows (no row
# at all in the Expert verdicts sheet), never "N/A"/0 placeholders.
# ---------------------------------------------------------------------------


def _write_batch_summary_sheet(
    ws: Worksheet, batch: BatchRecord, jobs: Sequence[BatchJobRecord]
) -> None:
    ws.title = "Summary"
    ws.append(["Atelier batch archive export"])
    ws.append(["Batch ID", batch.batch_id])
    ws.append(["Created (UTC)", batch.created_at])
    ws.append(["Updated (UTC)", batch.updated_at])
    ws.append(["Engine", batch.engine])
    ws.append(["Target source", batch.target_source])
    ws.append(["Target count", batch.target_count])
    ws.append(["Batch status", batch.status])
    ws.append([])

    status_counts: dict[str, int] = {}
    failure_counts: dict[str, int] = {}
    for job in jobs:
        status_counts[job.status] = status_counts.get(job.status, 0) + 1
        if job.failure is not None:
            failure_counts[job.failure.category] = failure_counts.get(job.failure.category, 0) + 1

    ws.append(["Job status", "Count"])
    for job_status, count in status_counts.items():
        ws.append([job_status, count])
    ws.append([])
    ws.append(["Failure category", "Count"])
    for category, count in failure_counts.items():
        ws.append([category, count])
    ws.append([])

    for note in batch.notes:
        ws.append(["Note", note])
    ws.append([])
    ws.append([_NON_PASS_FAIL_BANNER])


def _fmt_mode_list(value: object) -> str:
    if isinstance(value, list | tuple):
        return ", ".join(str(v) for v in value)
    return str(value) if value else ""


def _fmt_mode_counts(value: object) -> str:
    if isinstance(value, Mapping):
        return ", ".join(f"{k}={v}" for k, v in value.items())
    return str(value) if value else ""


def _write_jobs_sheet(ws: Worksheet, jobs: Sequence[BatchJobRecord]) -> None:
    ws.title = "Jobs"
    ws.append(
        [
            "job_id",
            "target_index",
            "target_label",
            "status",
            "degradation",
            "failure_category",
            "failure_message",
            "candidate_count",
            "ranked_count",
            "withheld_count",
            # BLOCKER-1 mode accounting: which modes the engine *asked* the
            # orchestrator to run, which actually produced candidates, and
            # the difference — so a Mode3-less "real" night reads as exactly
            # that in the morning workbook, never as a clean success column.
            "modes_requested",
            "modes_present",
            "missing_modes",
            "mode_counts",
            "candidate_set_pointer",
            "artifact_dir",
            "created_at",
            "updated_at",
        ]
    )
    for job in jobs:
        result_summary = job.result_summary or {}
        ws.append(
            [
                job.job_id,
                job.target_index,
                job.target_label,
                job.status,
                job.degradation or "",
                job.failure.category if job.failure else "",
                job.failure.message if job.failure else "",
                result_summary.get("candidate_count", ""),
                result_summary.get("ranked_count", ""),
                result_summary.get("withheld_count", ""),
                _fmt_mode_list(result_summary.get("modes_requested")),
                _fmt_mode_list(result_summary.get("modes_present")),
                _fmt_mode_list(result_summary.get("missing_modes")),
                _fmt_mode_counts(result_summary.get("mode_counts")),
                job.candidate_set_pointer or "",
                job.artifact_dir or "",
                job.created_at,
                job.updated_at,
            ]
        )


def _write_verdicts_sheet(ws: Worksheet, verdicts: Sequence[ExpertVerdict]) -> None:
    ws.title = "Expert verdicts"
    ws.append(["job_id", "candidate_key", "verdict_text", "reviewer", "recorded_at", "note"])
    for verdict in verdicts:
        ws.append(
            [
                verdict.job_id,
                verdict.candidate_key,
                verdict.verdict_text,
                verdict.reviewer,
                verdict.recorded_at,
                verdict.note or "",
            ]
        )


def build_batch_workbook(
    batch: BatchRecord,
    jobs: Sequence[BatchJobRecord],
    verdicts: Sequence[ExpertVerdict],
) -> bytes:
    """Three-sheet workbook: Summary (counts, no per-job detail), Jobs
    (machine-only columns), Expert verdicts (expert-only columns, one row
    per recorded verdict — jobs/candidates with no verdict simply have no
    row here, never a placeholder)."""
    wb = Workbook()
    summary_ws = wb.active
    assert summary_ws is not None
    _write_batch_summary_sheet(summary_ws, batch, jobs)
    _write_jobs_sheet(wb.create_sheet("Jobs"), jobs)
    _write_verdicts_sheet(wb.create_sheet("Expert verdicts"), verdicts)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
